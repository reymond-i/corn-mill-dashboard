import math
import hashlib
from io import BytesIO
from pathlib import Path
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
import openpyxl
import plotly.graph_objects as go

st.set_page_config(page_title="Corn Mill • Performance Dashboard", layout="wide")

def apply_theme():
    # A close match to the reference dashboard: light canvas, white cards, navy text, blue accents
    BG = "#F6F8FC"
    CARD = "#FFFFFF"
    NAVY = "#0B1F3A"
    MUTED = "#6B7280"
    BLUE = "#2563EB"
    BLUE_2 = "#1D4ED8"

    st.markdown(
        f"""
<style>
.stApp {{ background: {BG}; }}

section[data-testid="stSidebar"] {{
    background: {CARD};
    border-right: 1px solid rgba(15, 23, 42, 0.08);
}}
h1, h2, h3, h4 {{ color: {NAVY}; letter-spacing: -0.01em; }}
p, li, span {{ color: {NAVY}; }}

.stCaption, .small {{ color: {MUTED} !important; }}

.card {{
    background: {CARD};
    border: 1px solid rgba(15, 23, 42, 0.08);
    border-radius: 18px;
    padding: 14px 16px;
    box-shadow: 0 10px 28px rgba(15, 23, 42, 0.06);
}}

.kpi-title {{ color: {MUTED}; font-size: 12px; font-weight: 700; text-transform: uppercase; }}
.kpi-value {{ color: {NAVY}; font-size: 26px; font-weight: 900; line-height: 1.1; }}
.kpi-sub {{ color: {MUTED}; font-size: 12px; }}

.pill-pass {{
    background:#D4EDDA; color:#155724; padding:4px 10px; border-radius:999px; font-size:12px; font-weight:800;
}}
.pill-fail {{
    background:#F8D7DA; color:#721c24; padding:4px 10px; border-radius:999px; font-size:12px; font-weight:800;
}}

.stTabs [data-baseweb="tab-list"] {{ gap: 8px; }}
.stTabs [data-baseweb="tab"] {{
    background: rgba(37, 99, 235, 0.06);
    border: 1px solid rgba(37, 99, 235, 0.12);
    border-radius: 999px;
    padding: 10px 14px;
    color: {NAVY};
    font-weight: 700;
}}
.stTabs [aria-selected="true"] {{
    background: {BLUE};
    border-color: {BLUE_2};
    color: white !important;
}}

div[data-testid="stPlotlyChart"], div[data-testid="stDataFrame"] {{
    background: {CARD};
    border: 1px solid rgba(15, 23, 42, 0.08);
    border-radius: 18px;
    padding: 10px 12px;
    box-shadow: 0 10px 28px rgba(15, 23, 42, 0.05);
}}

.stButton button {{
    background: {BLUE};
    color: #FFFFFF;
    border-radius: 12px;
    border: 1px solid {BLUE_2};
    padding: 10px 14px;
    font-weight: 800;
}}
.stButton button:hover {{ background: {BLUE_2}; }}

hr {{ border-top: 1px solid rgba(15, 23, 42, 0.08); }}
</style>
""",
        unsafe_allow_html=True,
    )

apply_theme()

def is_nan(x):
    return isinstance(x, float) and math.isnan(x)

def to_float(x):
    if x is None or is_nan(x):
        return np.nan
    if isinstance(x, str):
        x = x.strip()
        if x in ["", "-"]:
            return np.nan
        try:
            return float(x)
        except:
            return np.nan
    try:
        return float(x)
    except:
        return np.nan

@st.cache_data(show_spinner=False)
def load_workbook(file_bytes: bytes, file_hash: str):
    # file_hash ensures Streamlit cache invalidates when you upload a different file
    bio = BytesIO(file_bytes)
    wb = openpyxl.load_workbook(bio, data_only=True)
    datasets = []

    def parse_sheet(sheet_name: str):
        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=None)
        ar = None
        for r in range(len(df)):
            if (df.iloc[r, :] == "Aperture, mm").any():
                ar = r
                break
        if ar is None:
            return []

        col_ap = int(np.where(df.iloc[ar, :] == "Aperture, mm")[0][0])

        def get_outlets(r):
            outs = []
            for c in range(col_ap + 1, df.shape[1]):
                v = df.iloc[r, c]
                if isinstance(v, str) and v.strip() and v.strip().lower() not in [
                    "percentage of grits collected at each outlet"
                ]:
                    outs.append(v.strip())
            return outs

        outlets = get_outlets(ar)
        if not outlets and ar + 1 < len(df):
            outlets = get_outlets(ar + 1)

        # dataset rows = numeric sample no in col 0
        data_rows = []
        for i, v in enumerate(df.iloc[:, 0].tolist()):
            if isinstance(v, (int, float)) and not is_nan(v):
                data_rows.append(i)
        data_rows = sorted(data_rows)

        out = []
        for k, start in enumerate(data_rows):
            end = data_rows[k + 1] if k + 1 < len(data_rows) else len(df)
            metrics = df.iloc[start, 0:9].tolist()

            sieve_rows = []
            for r in range(start, end):
                sieve = df.iloc[r, 9]
                ap = df.iloc[r, col_ap]
                if (not is_nan(ap) and ap is not None) or (isinstance(sieve, str) and sieve.strip()):
                    row = {"sieve": sieve, "aperture_mm": ap}
                    for j, o in enumerate(outlets):
                        row[o] = df.iloc[r, col_ap + 1 + j]
                    sieve_rows.append(row)

            out.append(
                dict(
                    sheet=sheet_name,
                    sample_no=int(metrics[0]) if not is_nan(metrics[0]) else None,
                    variety=metrics[1],
                    moisture_pct=metrics[2],
                    bulk_density_kgm3=metrics[3],
                    purity_pct=metrics[4],
                    input_kg=metrics[5],
                    main_product_recovery_pct=metrics[6],
                    byproduct_recovery_pct=metrics[7],
                    losses_pct=metrics[8],
                    outlets=outlets,
                    sieve_rows=sieve_rows,
                )
            )
        return out

    for s in wb.sheetnames:
        datasets.extend(parse_sheet(s))

    # ---- Build tables ----
    metrics_rows = []
    sieve_long = []

    for d in datasets:
        group_id = " | ".join(d["outlets"]) if d["outlets"] else "Unknown outlets"
        run_id = f"Run {d['sample_no']:02d}" if d["sample_no"] is not None else f"Sheet {d['sheet']}"

        metrics_rows.append(
            dict(
                Run=run_id,
                Sample_No=d["sample_no"],
                Sheet_Tab=d["sheet"],
                Outlet_Group=group_id,
                Variety=d["variety"],
                Moisture_Content_pct=to_float(d["moisture_pct"]),
                Bulk_Density_kgm3=to_float(d["bulk_density_kgm3"]),
                Purity_pct=to_float(d["purity_pct"]),
                Input_kg=to_float(d["input_kg"]),
                Main_Product_Recovery_pct=to_float(d["main_product_recovery_pct"]),
                Byproduct_Recovery_pct=to_float(d["byproduct_recovery_pct"]),
                Losses_pct=to_float(d["losses_pct"]),
            )
        )

        for row in d["sieve_rows"]:
            sieve = row["sieve"]
            ap = to_float(row["aperture_mm"])
            for o in d["outlets"]:
                sieve_long.append(
                    dict(
                        Run=run_id,
                        Sample_No=d["sample_no"],
                        Sheet_Tab=d["sheet"],
                        Outlet_Group=group_id,
                        Outlet=o,
                        Sieve=sieve,
                        Aperture_mm=ap,
                        Percent_Retained=to_float(row.get(o)),
                    )
                )

    metrics_df = pd.DataFrame(metrics_rows).sort_values("Sample_No").reset_index(drop=True)
    sieve_df = pd.DataFrame(sieve_long).dropna(subset=["Percent_Retained"]).copy()

    # PSD cumulative pass per run/outlet (exclude pan/NaN apertures)
    psd_rows = []
    for (run, outlet), g in sieve_df.groupby(["Run", "Outlet"]):
        g2 = g.dropna(subset=["Aperture_mm"]).copy()
        if g2.empty:
            continue
        g2 = g2.sort_values("Aperture_mm", ascending=False)
        g2["Cum_Retained"] = g2["Percent_Retained"].cumsum()
        g2["Cum_Pass_pct"] = 100 - g2["Cum_Retained"]
        for _, r in g2.iterrows():
            psd_rows.append(
                dict(
                    Run=run,
                    Outlet=outlet,
                    Aperture_mm=float(r["Aperture_mm"]),
                    Cum_Pass_pct=float(r["Cum_Pass_pct"]),
                )
            )
    psd_df = pd.DataFrame(psd_rows)

    # ---- Year-based compliance (as requested) ----
    def compute_standard(sample_no: int) -> str:
        if sample_no is None or pd.isna(sample_no):
            return "Unknown"
        return "2018" if int(sample_no) <= 8 else "2021"

    metrics_df["Standard_Used"] = metrics_df["Sample_No"].apply(compute_standard)

    def compliant(row) -> bool:
        loss_ok = row["Losses_pct"] <= 5
        if row["Standard_Used"] == "2018":
            main_ok = row["Main_Product_Recovery_pct"] >= 64
        elif row["Standard_Used"] == "2021":
            main_ok = row["Main_Product_Recovery_pct"] >= 55
        else:
            return False
        return bool(main_ok and loss_ok)

    metrics_df["Compliance"] = metrics_df.apply(lambda r: "PASS" if compliant(r) else "FAIL", axis=1)

    def fail_reason(row) -> str:
        loss_fail = row["Losses_pct"] > 5
        if row["Standard_Used"] == "2018":
            rec_fail = row["Main_Product_Recovery_pct"] < 64
        elif row["Standard_Used"] == "2021":
            rec_fail = row["Main_Product_Recovery_pct"] < 55
        else:
            rec_fail = True
        if rec_fail and loss_fail:
            return "Recovery + Losses"
        if rec_fail:
            return "Recovery"
        if loss_fail:
            return "Losses"
        return "—"
    metrics_df["Fail_Reason"] = metrics_df.apply(fail_reason, axis=1)

    # Batch-level compliance score (0–100)
    def compliance_score(row) -> float:
        if row["Standard_Used"] == "2018":
            rec_thr = 64.0
        else:
            rec_thr = 55.0
        rec = row["Main_Product_Recovery_pct"]
        loss = row["Losses_pct"]
        if pd.isna(rec) or pd.isna(loss):
            return float("nan")
        rec_ratio = max(0.0, min(1.0, float(rec) / rec_thr))
        loss_ratio = max(0.0, min(1.0, (5.0 - float(loss)) / 5.0))
        return 100.0 * (0.5 * rec_ratio + 0.5 * loss_ratio)

    metrics_df["Compliance_Score_%"] = metrics_df.apply(compliance_score, axis=1)

    return metrics_df, sieve_df, psd_df

def kpi_card(title: str, value: str, subtitle: str = ""):
    st.markdown(
        f"""
<div class="card">
  <div class="kpi-title">{title}</div>
  <div class="kpi-value">{value}</div>
  <div class="kpi-sub">{subtitle}</div>
</div>
""",
        unsafe_allow_html=True,
    )

def pill(text: str, ok: bool):
    cls = "pill-pass" if ok else "pill-fail"
    st.markdown(f"<span class='{cls}'>{text}</span>", unsafe_allow_html=True)

def fig_psd_single_run(psd_df: pd.DataFrame, run: str):
    d = psd_df[psd_df["Run"] == run].copy()
    fig = go.Figure()
    for outlet in sorted(d["Outlet"].unique().tolist()):
        g = d[d["Outlet"] == outlet].sort_values("Aperture_mm", ascending=True)
        fig.add_trace(
            go.Scatter(
                x=g["Aperture_mm"],
                y=g["Cum_Pass_pct"],
                mode="lines+markers",
                line_shape="spline",
                marker=dict(symbol="circle", size=9, line=dict(width=1, color="black")),
                line=dict(width=2),
                name=outlet,
            )
        )
    fig.update_layout(
        template="plotly_white",
        title=f"Particle Size Distribution Curve — {run}",
        xaxis=dict(title="Aperture (mm) — log scale", type="log"),
        yaxis=dict(title="Cumulative % Passing", range=[0, 100]),
        height=480,
        margin=dict(l=60, r=20, t=60, b=50),
        legend_title="Outlet",
        font=dict(family="system-ui", size=12),
    )
    return fig

def fig_bar_sieve_all_outlets(sieve_df: pd.DataFrame, run: str):
    # One chart per run: compare ALL outlets in the same bar chart
    d = sieve_df[(sieve_df["Run"] == run)].dropna(subset=["Aperture_mm"]).copy()
    if d.empty:
        return go.Figure()
    d["Sieve_Label"] = d.apply(lambda r: f"{r['Sieve']} ({r['Aperture_mm']:g} mm)", axis=1)
    # Order by aperture descending
    order = (
        d[["Sieve_Label", "Aperture_mm"]]
        .drop_duplicates()
        .sort_values("Aperture_mm", ascending=False)["Sieve_Label"]
        .tolist()
    )
    piv = d.pivot_table(index="Sieve_Label", columns="Outlet", values="Percent_Retained", aggfunc="mean").reindex(order)
    fig = go.Figure()
    for outlet in piv.columns.tolist():
        fig.add_trace(go.Bar(x=piv.index.tolist(), y=piv[outlet].fillna(0).tolist(), name=outlet))
    fig.update_layout(
        template="plotly_white",
        title=f"Sieve Distribution — Compare Outlets (Percent Retained) — {run}",
        xaxis=dict(title="Sieve (Aperture mm)", tickangle=30),
        yaxis=dict(title="Percent Retained (%)"),
        barmode="group",
        height=520,
        margin=dict(l=60, r=20, t=60, b=140),
        font=dict(family="system-ui", size=12),
        legend_title="Outlet",
    )
    return fig

def fig_compliance_pie(metrics_df: pd.DataFrame):
    counts = metrics_df["Compliance"].value_counts().reindex(["PASS", "FAIL"]).fillna(0).astype(int)
    fig = go.Figure(data=[go.Pie(labels=counts.index.tolist(), values=counts.values.tolist(), hole=0.55)])
    fig.update_layout(
        template="plotly_white",
        title="Compliance Summary (PASS vs FAIL)",
        height=320,
        margin=dict(l=20, r=20, t=60, b=10),
        font=dict(family="system-ui", size=12),
    )
    return fig

def year_group_label(sample_no: int) -> str:
    return "2018 group (Run 01–08)" if sample_no <= 8 else "2021 group (Run 09+)"

# ------------------ Upload (required) ------------------
with st.sidebar:
    st.markdown("### University of the Philippines Los Baños")
    st.markdown("**Corn mill • performance dashboard**")
    st.write("Upload the Excel workbook (.xlsx) to generate the dashboard.")
    up = st.file_uploader("Upload .xlsx", type=["xlsx"], key="uploader")
    if up is None:
        st.info("Please upload the Excel file to start.")
        st.stop()
    file_bytes = up.getvalue()
    file_hash = hashlib.md5(file_bytes).hexdigest()

metrics_df, sieve_df, psd_df = load_workbook(file_bytes, file_hash)

# ---- Year filter toggle ----
with st.sidebar:
    st.markdown("---")
    st.markdown("### Filters")
    year_choice = st.radio(
        "Show runs from:",
        options=["All (2018 + 2021)", "2018 group (Run 01–08)", "2021 group (Run 09+)"],
        index=0,
    )

if year_choice == "2018 group (Run 01–08)":
    metrics_f = metrics_df[metrics_df["Sample_No"] <= 8].copy()
elif year_choice == "2021 group (Run 09+)":
    metrics_f = metrics_df[metrics_df["Sample_No"] >= 9].copy()
else:
    metrics_f = metrics_df.copy()

runs_f = metrics_f["Run"].tolist()
if not runs_f:
    st.warning("No runs match the selected year filter.")
    st.stop()

sieve_f = sieve_df[sieve_df["Run"].isin(runs_f)].copy()
psd_f = psd_df[psd_df["Run"].isin(runs_f)].copy()

groups_f = sorted(metrics_f["Outlet_Group"].unique().tolist())

with st.sidebar:
    group_sel = st.selectbox("Outlet group", options=groups_f)
    runs_in_group = metrics_f.loc[metrics_f["Outlet_Group"] == group_sel, "Run"].tolist()
    run_sel = st.selectbox("Run", options=runs_in_group)

# Selected row (for KPIs)
row = metrics_df[metrics_df["Run"] == run_sel].iloc[0]
std = row["Standard_Used"]
year_group = year_group_label(int(row["Sample_No"]))

# Derived KPIs similar to the reference layout
total_milling_recovery = row["Main_Product_Recovery_pct"] + row["Byproduct_Recovery_pct"] if not pd.isna(row["Byproduct_Recovery_pct"]) else row["Main_Product_Recovery_pct"]
losses = row["Losses_pct"]
capacity = row["Input_kg"]

tabs = st.tabs(["Dashboard", "Particle Distribution", "Sieve Distribution", "Library", "Team"])

# ------------------ DASHBOARD TAB ------------------
with tabs[0]:
    st.markdown("# Dashboard")
    st.caption(datetime.now().strftime("%A %B %d, %Y %I:%M %p"))

    c1, c2, c3 = st.columns(3)
    with c1:
        kpi_card("Total Milling Recovery", f"{total_milling_recovery:.2f}%", "Main + by-product recovery")
    with c2:
        kpi_card("Losses", f"{losses:.2f}%", "Standard limit: ≤ 5%")
    with c3:
        kpi_card("Milling Capacity, kg", f"{capacity:.2f}", "Input per run")

    st.markdown("### Particle Size Distribution Curve")
    left, right = st.columns([1.5, 1.0])
    with left:
        st.plotly_chart(fig_psd_single_run(psd_f, run_sel), use_container_width=True)
    with right:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("#### Your corn mill performance in one glance.")
        st.write("Monitoring the performance of your corn mill has now become more accessible and easier on the eye.")
        st.markdown("---")
        st.write("**Moisture Content**")
        kpi_card("Moisture Content", f"{row['Moisture_Content_pct']:.2f}%", "")
        st.write("**Purity**")
        kpi_card("Purity", f"{row['Purity_pct']:.2f}%", "")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("### Compliance Summary")
    pass_n = int((metrics_f["Compliance"] == "PASS").sum())
    total_n = int(len(metrics_f))
    fail_n = total_n - pass_n
    avg_score = float(metrics_f["Compliance_Score_%"].mean())
    p1, p2 = st.columns([1.0, 1.4])
    with p1:
        st.plotly_chart(fig_compliance_pie(metrics_f), use_container_width=True)
    with p2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.write(f"**Year filter:** {year_choice}")
        st.write(f"Runs shown: **{total_n}**")
        st.write(f"PASS: **{pass_n}**  |  FAIL: **{fail_n}**")
        st.write(f"Average compliance score: **{avg_score:.1f}%**")
        st.write("")
        st.write("Runs that **PASSED** (sorted by score)")
        passed = metrics_f[metrics_f["Compliance"] == "PASS"].copy()
        if passed.empty:
            st.info("No runs passed under the selected filter.")
        else:
            passed_sorted = passed.sort_values("Compliance_Score_%", ascending=False)[
                ["Run","Compliance_Score_%","Standard_Used"]
            ]
            st.dataframe(passed_sorted, use_container_width=True, hide_index=True)

        st.write("Failed runs (with reason)")
        failed = metrics_f[metrics_f["Compliance"] == "FAIL"].copy()
        if failed.empty:
            st.info("No failed runs under the selected filter.")
        else:
            failed_sorted = failed.sort_values("Compliance_Score_%", ascending=True)[
                ["Run","Compliance_Score_%","Standard_Used","Fail_Reason","Main_Product_Recovery_pct","Losses_pct"]
            ].rename(columns={
                "Main_Product_Recovery_pct":"Main Recovery (%)",
                "Losses_pct":"Losses (%)"
            })
            st.dataframe(failed_sorted, use_container_width=True, hide_index=True)
        st.markdown('</div>', unsafe_allow_html=True)

    st.caption(f"Compliance basis: {year_group} uses {std} thresholds for Main Product Recovery + Losses.")

# ------------------ PARTICLE DISTRIBUTION TAB ------------------
with tabs[1]:
    st.markdown("# Particle Size Distribution")
    st.caption("Semi-log aperture (mm) vs cumulative % passing, with circle markers per data point.")

    # Mimic outlet tabs (#6-8, #10-12, etc.) by using outlet groups
    group_tabs = st.tabs(groups_f if groups_f else ["(no groups found)"])
    for i, gname in enumerate(groups_f):
        with group_tabs[i]:
            runs_g = metrics_f.loc[metrics_f["Outlet_Group"] == gname, "Run"].tolist()
            run_g = st.selectbox("Select run", options=runs_g, key=f"psd_run_{i}")
            st.plotly_chart(fig_psd_single_run(psd_f, run_g), use_container_width=True)

# ------------------ SIEVE DISTRIBUTION TAB ------------------
with tabs[2]:
    st.markdown("# Sieve Distribution")
    st.caption("One bar chart per run: all outlets compared in the same chart (grouped bars).")
    st.plotly_chart(fig_bar_sieve_all_outlets(sieve_f, run_sel), use_container_width=True)

# ------------------ LIBRARY TAB ------------------
with tabs[3]:
    st.markdown("# Library")
    st.write("This app is upload-based (no permanent storage).")
    st.write(f"Current file: **{up.name}**")
    st.download_button("Download uploaded file", data=file_bytes, file_name=up.name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.write("Recent file history is not persisted on Streamlit Cloud in free tier.")

# ------------------ TEAM TAB ------------------
with tabs[4]:
    st.markdown("# Team")
    st.write("Corn Group")
    cols = st.columns(4)
    members = ["Janreich", "Raymond", "Micko", "Reyniel"]
    for c, m in zip(cols, members):
        with c:
            st.markdown(f"<div class='card'><div class='kpi-title'>Member</div><div class='kpi-value' style='font-size:20px'>{m}</div></div>", unsafe_allow_html=True)
